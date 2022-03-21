# -*- coding: utf-8 -*-

import os
import sys
import argparse
import re
from openpyxl import load_workbook
from openpyxl import Workbook


def DebugLog(*msgs):
    if True:
        print('[DEBUG]: ', msgs)


def GetArgs():
    parser = argparse.ArgumentParser(description='Delete duplicated lines for excel.')

    parser.add_argument('--file', dest='file', required=True)
    parser.add_argument('--sheet', dest='sheet', required=True)
    parser.add_argument('--column', dest='column', required=True)
    args = parser.parse_args()
    return args
    
def SortByName(t):
    return t[0]

def Run(file, sheet, column):
    rb = load_workbook(file)
    DebugLog(rb.sheetnames)
    r_sheet = rb[sheet]
    DebugLog(r_sheet.title)
    result_dict = {}
    result_list = []
    for row in r_sheet.rows:
        name = row[column].value
        if name:
            name = name.replace(' ', '').strip().lower()
            DebugLog('Name: ', name)
            result_dict[name] = row
            result_list.append((name, row))
    DebugLog('Original row number: ', len(result_list) )
    DebugLog('No Duplicated row number: ', len(result_dict) )
    DebugLog('Duplicated: ', len(result_list) -  len(result_dict))
    result_list.sort(key=SortByName)
    # for name, row in result_dict.items():
    #     print(name) 
    rb.close()

    wb = Workbook()
    w_sheet = wb.active    

    row_num = 1
    for name, rows in result_dict.items():
        w_sheet.insert_rows(idx=row_num)
        col_num = 1    
        for cell in rows:
            w_sheet.cell(row=row_num, column=col_num).value = cell.value
            col_num += 1
        row_num += 1

    new_file_name = file[:-5] + '_no_duplicated.xlsx'
    DebugLog('New file name: ', new_file_name)
    wb.save(filename=new_file_name)
    wb.close()


if __name__ == '__main__':
    args = GetArgs()
    if args.file and args.sheet and args.column:
        Run(args.file, args.sheet, int(args.column))
    
    sys.exit(0)
